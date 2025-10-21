#!/usr/bin/env python3

import openpyxl
import pandas as pd
import random
import os
from openpyxl.styles import Font

# ✅ 파일 로드 (엑셀)
file_path = 'Auto_Table.xlsx'

# ✅ Participants 시트에서 참가자 명단 읽기
df_participants = pd.read_excel(file_path, sheet_name='Participants')
male_players = df_participants['남자'].dropna().tolist()
female_players = df_participants['여자'].dropna().tolist()

num_male = len(male_players)
num_female = len(female_players)

print(f"참가자: 남 {num_male}, 여 {num_female}")

# ✅ 남여 비율별 시퀀스 표

# 혼복 6
sequence_table = {
    # 20명 (5코트)
    (6,14): [2,2,2,13,13], (7,13): [7,7,7,7,10], (8,12): [2,3,3,13,13],
    (9,11): [3,3,3,13,13], (10,10): [7,7,8,8,11], (11,9): [3,3,8,8,15],
    (12,8): [3,8,8,8,14], (13,7): [4,8,8,8,14], (14,6): [4,4,8,12,14],
    (15,5): [4,4,12,12,12], (16,4): [4,9,9,12,12],
    
    # 15명 (3코트)
    (2,13): [16,16,16,18,19], (3,12): [16,16,17,18,19], (4,11): [16,17,17,18,18],
    (5,10): [17,17,17,18,18], (6,9): [17,17,18,18,21], (7,8): [17,18,18,21,21],
    (8,7): [18,18,21,21,21], (9,6): [21,21,18,21,18], (10,5): [21,21,22,18,18],
    (11,4): [18,22,22,22,24], (12,3): [19,22,22,24,24], (13,2): [19,22,24,24,24],
    
    # 10명 (2코트)
    (2,8): [26,26,26,27,27], (3,7): [26,26,27,27,27], (4,6): [26,27,27,27,28],
    (5,5): [27,27,27,28,28], (6,4): [27,27,28,28,30], (7,3): [27,28,30,30,31],
    (8,2): [28,30,31,31,31],
}

# 혼복 8
sequence_table_v2 = {
    # 20명 (5코트)
    (6,14): [2,7,10,10,13], (7,13): [7,7,7,10,13], (8,12): [3,10,10,11,11],
    (9,11): [3,3,7,13,15], (10,10): [3,7,8,11,15], (11,9): [8,8,11,11,11],
    (12,8): [8,8,8,8,15], (13,7): [8,8,8,12,14], (14,6): [4,8,12,12,14],
    (15,5): [4,12,12,12,12], (16,4): [5,12,12,12,12],
    
    # 15명 (3코트)
    (2,13): [16,16,17,17,18], (3,12): [16,17,17,17,18], (4,11): [17,17,17,17,18],
    (5,10): [17,17,17,18,18], (6,9): [17,17,18,18,21], (7,8): [17,18,21,21,21],
    (8,7): [18,21,21,21,22], (9,6): [21,21,18,21,18], (10,5): [22,18,22,18,18],
    (11,4): [18,22,22,22,24], (12,3): [19,22,22,24,24], (13,2): [19,22,24,24,24],
    
    # 10명 (2코트)
    (2,8): [26,26,26,27,27], (3,7): [26,26,27,27,27], (4,6): [27,27,27,27,27],
    (5,5): [27,27,27,28,30], (6,4): [27,27,30,30,31], (7,3): [27,30,30,31,31],
    (8,2): [30,30,31,31,31],
}

use_v2 = True  # ✅ True면 혼복8, False면 혼복6

if use_v2:
    round_combinations = sequence_table_v2.get((num_male, num_female), None)
else:
    round_combinations = sequence_table.get((num_male, num_female), None)

if round_combinations is None:
    raise ValueError(f"⚠️ 현재 참가자 수 (남 {num_male}, 여 {num_female}) 에 대한 시퀀스가 정의되어 있지 않습니다.")
print(f"사용 시퀀스: {round_combinations}")

wb = openpyxl.load_workbook(file_path)
ws = wb['Match_schedule']

# ============================================================================
# 라이프 멤버 명단 (여기만 수정하기!)
# ============================================================================
life_members_male = [
    '김종현', '문광식', '박동언', '박종성', '오성목', '임채경', '정기완', '조창현', '홍상현'
]
life_members_female = [
    '김예인', '문지정', '서가연', '서자랑', '장은비', '정예원', '최은진'
]
# ============================================================================

life_members = life_members_male + life_members_female

# 엑셀에서 그룹 정보 로드 (A그룹/B그룹)
df_life = pd.read_excel(file_path, sheet_name='LIFE_members', header=None, skiprows=2)
group_dict = {}
for name in df_life.iloc[:,1].dropna():  # B열: A그룹 남자
    group_dict[name] = 'A'
for name in df_life.iloc[:,2].dropna():  # C열: B그룹 남자
    group_dict[name] = 'B'
for name in df_life.iloc[:,3].dropna():  # D열: A그룹 여자
    group_dict[name] = 'A'
for name in df_life.iloc[:,4].dropna():  # E열: B그룹 여자
    group_dict[name] = 'B'

# ✅ 게스트 선수들을 group_dict 에 'guest' 로 등록
for p in male_players + female_players:
    if p not in group_dict:
        group_dict[p] = 'guest'

combi_table = {
    # 20명 (5코트, 4경기)
    1: (0,0,4), 2: (0,1,3), 3: (0,2,2), 4: (0,3,1), 5: (0,4,0),
    6: (1,0,3), 7: (1,1,2), 8: (1,2,1), 9: (1,3,0), 10: (2,0,2),
    11: (2,1,1), 12: (2,2,0), 13: (3,0,1), 14: (3,1,0), 15: (4,0,0),
    
    # 15명 (3코트, 3경기)
    16: (0,0,3), 17: (0,1,2), 18: (0,2,1), 19: (0,3,0),
    20: (1,0,2), 21: (1,1,1), 22: (1,2,0),
    23: (2,0,1), 24: (2,1,0), 25: (3,0,0),
    
    # 10명 (2코트, 2경기)
    26: (0,0,2), 27: (0,1,1), 28: (0,2,0),
    29: (1,0,1), 30: (1,1,0), 31: (2,0,0),
}

round_rows = [5,7,9,11,13]
all_players = male_players + female_players

# ✅ swap_if_needed 함수 (cross-pair swap)
def swap_if_needed(previous_round, current_round, max_attempts=20):
    attempt = 0
    swap_warning = False
    while attempt < max_attempts:
        need_retry = False
        for prev_team in previous_round:
            for curr_team in current_round:
                prev_players = prev_team[1]
                curr_players = curr_team[1]
                common_players = set(map(tuple, prev_players)) & set(map(tuple, curr_players))
                if len(common_players) >= 3 and len(curr_players) >= 4:
                    p1, p2, p3, p4 = curr_players[:4]
                    curr_players[0] = p1
                    curr_players[1] = p3
                    curr_players[2] = p2
                    curr_players[3] = p4
                    need_retry = True
        if not need_retry:
            break
        attempt += 1
    if attempt >= max_attempts:
        swap_warning = True
    return current_round, swap_warning


MAX_TRIALS = 100
trial = 0
all_rounds_matches = []  # Store all matches from all rounds
while trial < MAX_TRIALS:
    trial += 1
    # ✅ 루프 시작 시 변수 초기화
    rest_count = {p:0 for p in all_players}
    game_count = {p:0 for p in all_players}
    mixed_played_men = {p:0 for p in male_players}
    mixed_played_women = {p:0 for p in female_players}
    previous_round = []
    swap_warning = False
    all_rounds_matches = []  # Reset for each trial

    # ✅ main loop
    for rnd, comb_num in enumerate(round_combinations):
        mixed, men_d, women_d = combi_table[comb_num]
        men_need = mixed*2 + men_d*4
        women_need = mixed*2 + women_d*4

        rest_num = len(all_players) - (men_need + women_need)
        rest_this_round = []
        active_men = male_players.copy()
        active_women = female_players.copy()
        
        # ✅ 휴식자 선정: 휴식 횟수가 가장 적은 사람부터 우선 배정
        # 먼저 최소 휴식 횟수를 찾음
        min_rest = min(rest_count.values())
        
        # 최소 휴식 횟수인 사람들을 우선 휴식 배정
        never_or_least_rested = [p for p in all_players if rest_count[p] == min_rest]
        
        for p in never_or_least_rested:
            if len(rest_this_round) >= rest_num:
                break
            
            # 성별 체크: 해당 성별이 경기에 충분한지 확인
            if p in active_men:
                remaining_men = len([x for x in active_men if x not in rest_this_round]) - 1
                if remaining_men >= men_need:
                    rest_this_round.append(p)
                    rest_count[p] += 1
                    active_men.remove(p)
            elif p in active_women:
                remaining_women = len([x for x in active_women if x not in rest_this_round]) - 1
                if remaining_women >= women_need:
                    rest_this_round.append(p)
                    rest_count[p] += 1
                    active_women.remove(p)
        
        # ✅ 추가 휴식자 선정 (아직 슬롯이 남았다면)
        while len(rest_this_round) < rest_num:
            # 휴식 가능한 남자/여자 후보 필터링 (휴식 횟수 순으로 정렬)
            men_candidates = [p for p in active_men if p not in rest_this_round and len(active_men) - len([x for x in rest_this_round if x in active_men]) > men_need]
            women_candidates = [p for p in active_women if p not in rest_this_round and len(active_women) - len([x for x in rest_this_round if x in active_women]) > women_need]
            
            # 휴식 횟수로 정렬 (적은 순)
            men_candidates.sort(key=lambda x: rest_count[x])
            women_candidates.sort(key=lambda x: rest_count[x])
            
            candidates = men_candidates + women_candidates
            
            if not candidates:
                # 조건을 만족하는 후보가 없으면 재시도 (Trial 실패 조건)
                break
            
            # 가장 휴식이 적은 사람 선택
            candidates.sort(key=lambda x: rest_count[x])
            p = candidates[0]
            
            if p in active_men:
                active_men.remove(p)
                rest_this_round.append(p)
                rest_count[p] += 1
            elif p in active_women:
                active_women.remove(p)
                rest_this_round.append(p)
                rest_count[p] += 1

        match_list = []

        # ✅ 혼복 (모든 선수 최소 1회 참여 우선)
        for _ in range(mixed):
            unplayed_men = [p for p in active_men if mixed_played_men[p]==0]
            unplayed_women = [p for p in active_women if mixed_played_women[p]==0]

            men_pool = unplayed_men if len(unplayed_men) >= 2 else active_men
            women_pool = unplayed_women if len(unplayed_women) >= 2 else active_women

            men_pool.sort(key=lambda x: game_count[x])
            women_pool.sort(key=lambda x: game_count[x])
            random.shuffle(men_pool)
            random.shuffle(women_pool)

            men_pair = men_pool[:2]
            women_pair = women_pool[:2]

            for p in men_pair:
                mixed_played_men[p] += 1
            for p in women_pair:
                mixed_played_women[p] += 1

            team = [men_pair[0], women_pair[0], men_pair[1], women_pair[1]]
            match_list.append(('혼복', team))

            for p in men_pair + women_pair:
                game_count[p] += 1
            active_men = [m for m in active_men if m not in men_pair]
            active_women = [w for w in active_women if w not in women_pair]

        # ✅ 남복 (A/B 그룹 우선)
        for _ in range(men_d):
            active_men.sort(key=lambda x: game_count[x])
            random.shuffle(active_men)
            a_men = [p for p in active_men if group_dict.get(p) == 'A']
            guest_men = [p for p in active_men if group_dict.get(p) == 'guest']
            if len(a_men) + len(guest_men) >= 4 and len(a_men) >= 1:
                combined = a_men + guest_men
                team = combined[:4]
            else:
                b_men = [p for p in active_men if group_dict.get(p) == 'B']
                if len(b_men) + len(guest_men) >= 4 and len(b_men) >= 1:
                    combined = b_men + guest_men
                    team = combined[:4]
                else:
                    team = active_men[:4]

            match_list.append(('남복', team))
            for p in team:
                game_count[p] += 1
            active_men = [m for m in active_men if m not in team]

        # ✅ 여복 (A/B 그룹 우선)
        for _ in range(women_d):
            active_women.sort(key=lambda x: game_count[x])
            random.shuffle(active_women)
            a_women = [p for p in active_women if group_dict.get(p) == 'A']
            guest_women = [p for p in active_women if group_dict.get(p) == 'guest']
            if len(a_women) + len(guest_women) >= 4 and len(a_women) >= 1:
                combined = a_women + guest_women
                team = combined[:4]
            else:
                b_women = [p for p in active_women if group_dict.get(p) == 'B']
                if len(b_women) + len(guest_women) >= 4 and len(b_women) >= 1:
                    combined = b_women + guest_women
                    team = combined[:4]
                else:
                    team = active_women[:4]

            match_list.append(('여복', team))
            for p in team:
                game_count[p] += 1
            active_women = [w for w in active_women if w not in team]

        # ✅ swap 로직 적용
        match_list, swap_flag = swap_if_needed(previous_round, match_list)
        if swap_flag:
            swap_warning = True
        previous_round = match_list.copy()

        # 각 선수의 이름에 성별 식별자 추가
        for i, team in enumerate(match_list):
            players_with_gender = [f"{p}(m)" if p in male_players else f"{p}(f)" for p in team[1]]
            match_list[i] = (team[0], players_with_gender)

        # Select leader from life members after gender identifiers are added
        for match_index, match in enumerate(match_list):
            team = match[1]
            # Extract name without gender identifier and check if it's a life member
            life_members_in_team = [p for p in team if p.split('(')[0] in life_members]
            if life_members_in_team:
                leader = life_members_in_team[0]  # Select the first life member as leader
                # Reorder team with leader first
                new_team = [leader] + [p for p in team if p != leader]
                match_list[match_index] = (match[0], new_team)
        
        # ✅ 코트 배정 (리더 선정 후)
        match_list_sorted = sorted(match_list, key=lambda x: 0 if x[0]=='여복' else 1)
        match_players_with_leaders = []
        for m in match_list_sorted:
            match_players_with_leaders.extend(m[1])
        
        # 코트 수에 따라 빈 코트 처리
        num_courts = len(match_list_sorted)  # 실제 사용 코트 수
        empty_slots = max(0, (4 - num_courts) * 4)  # 빈 코트 슬롯 (4코트 기준, 5번째는 휴식)
        
        # 휴식자에도 성별 태그 추가
        rest_with_gender = [f"{p}(m)" if p in male_players else f"{p}(f)" for p in rest_this_round]
        
        if rnd == 0:  # 첫 라운드만 디버깅 출력
            print(f"\n[디버깅] 라운드 {rnd+1} 코트 배정:")
            print(f"  경기 수: {num_courts}개")
            print(f"  경기 플레이어: {len(match_players_with_leaders)}명")
            print(f"  빈 슬롯: {empty_slots}개")
            print(f"  휴식자: {len(rest_with_gender)}명")
            print(f"  총 길이: {len(match_players_with_leaders)} + {empty_slots} + {len(rest_with_gender)} = {len(match_players_with_leaders) + empty_slots + len(rest_with_gender)}")
        
        # 경기 플레이어 + 빈 코트 + 휴식자
        final_players = match_players_with_leaders + ([None] * empty_slots) + rest_with_gender
        final_players = final_players[:20] + [None]*(20-len(final_players))

        row = round_rows[rnd]
        # dd_mooon : 엑셀 파일에 저장할 때 라이프 멤버는 * 표시
        for idx, name in enumerate(final_players):
            if name:
                cell = ws.cell(row=row, column=idx+3)
                cell.value = name
                name_without_gender = name.split('(')[0]  # Extract name without gender
                if name_without_gender in life_members:
                    cell.value = f"*{name}"
                # Set background color based on team composition
                if name.endswith('(f)'):
                    # Count females in the match
                    pass  # Will handle color below
            else:
                ws.cell(row=row, column=idx+3).value = None
        
        # Store matches for this round
        all_rounds_matches.extend([(rnd + 1, match[0], match[1]) for match in match_list])

    # ✅ 혼복 최소 1회 미참여 선수 확인
    unplayed_men_final = [p for p,v in mixed_played_men.items() if v==0]
    unplayed_women_final = [p for p,v in mixed_played_women.items() if v==0]
    
    # 휴식 0번인 사람 확인
    never_rested = [p for p in all_players if rest_count[p] == 0]

    print(f"Trial {trial}: 미혼복 남={len(unplayed_men_final)}, 여={len(unplayed_women_final)}, swap_warning={swap_warning}, 미휴식={len(never_rested)}")
    
    if trial <= 3 or trial >= 98:  # 처음 3번과 마지막 2번만 상세 출력
        print(f"  상세: swap_warning={swap_warning}, never_rested={never_rested}")
        print(f"  휴식 카운트: {[(p, rest_count[p]) for p in all_players]}")

    # 성공 조건: 인원수에 따라 다르게
    total_players = len(all_players)
    
    if total_players == 20:
        # 20명: 혼복도 중요하게, 휴식도 확인
        if len(never_rested) == 0 and (len(unplayed_men_final) + len(unplayed_women_final)) <= 2:
            print("✅ 성공적으로 매칭 완료 (20명)")
            break
    else:
        # 15명 이하: 휴식만 중요, 혼복은 무시
        if len(never_rested) == 0:
            print(f"✅ 성공적으로 매칭 완료 ({total_players}명)")
            print(f"  (혼복 미참여 {len(unplayed_men_final) + len(unplayed_women_final)}명 있지만 허용)")
            break

# ✅ 파일명 자동 증가 저장
base_filename = 'LIFE_Auto_Table'
file_ext = '.xlsx'
file_path_save = base_filename + file_ext
counter = 2
while os.path.exists(file_path_save):
    file_path_save = f"{base_filename}_{counter}{file_ext}"
    counter += 1

wb.save(file_path_save)

# dd_mooon : 통계 데이터 수집
player_stats = {}
for player in all_players:
    player_stats[player] = {
        '혼복': 0,
        '남복': 0,
        '여복': 0,
        '총게임': 0,
        '휴식': 0
    }

# dd_mooon : 각 라운드별로 통계 계산
total_rounds = len(all_rounds_matches) // (len(set([r[0] for r in all_rounds_matches])))
rounds_count = max([r[0] for r in all_rounds_matches])

for round_num, match_type, team in all_rounds_matches:
    for player_with_gender in team:
        player_name = player_with_gender.split('(')[0]
        if player_name in player_stats:
            player_stats[player_name][match_type] += 1
            player_stats[player_name]['총게임'] += 1

# dd_mooon : 휴식 횟수 계산
for player in all_players:
    player_stats[player]['휴식'] = rounds_count - player_stats[player]['총게임']

# dd_mooon : 테이블 스타일링
def pad_korean(text, width):
    """한글은 2칸, 영문/숫자는 1칸으로 계산하여 패딩"""
    text_width = sum(2 if ord(c) > 127 else 1 for c in text)
    padding = width - text_width
    return text + ' ' * max(0, padding)

# dd_mooon : 각 플레이어별 통계 출력
print("\n" + "="*80)
print("게임 통계")
print("="*80)

# dd_mooon : 전체 플레이어 통합 출력 (성별 컬럼 추가)
print(f"{pad_korean('이름', 15)} {pad_korean('성별', 6)} {pad_korean('구분', 12)} {pad_korean('총게임', 8)} {pad_korean('혼복', 6)} {pad_korean('남복', 6)} {pad_korean('여복', 6)} {pad_korean('휴식', 6)}")
print("-" * 80)
for player in male_players:
    member_type = "라이프" if player in life_members else "게스트"
    stats = player_stats[player]
    print(f"{pad_korean(player, 15)} {pad_korean('남', 6)} {pad_korean(member_type, 12)} {pad_korean(str(stats['총게임']), 8)} {pad_korean(str(stats['혼복']), 6)} {pad_korean(str(stats['남복']), 6)} {pad_korean(str(stats['여복']), 6)} {pad_korean(str(stats['휴식']), 6)}")

for player in female_players:
    member_type = "라이프" if player in life_members else "게스트"
    stats = player_stats[player]
    print(f"{pad_korean(player, 15)} {pad_korean('여', 6)} {pad_korean(member_type, 12)} {pad_korean(str(stats['총게임']), 8)} {pad_korean(str(stats['혼복']), 6)} {pad_korean(str(stats['남복']), 6)} {pad_korean(str(stats['여복']), 6)} {pad_korean(str(stats['휴식']), 6)}")

print("\n" + "="*80)
print(f"✅ 저장 완료: {file_path_save}")
print("="*80)
print()
print("🎾🎾 테니스 치러 가요~ 🎾🎾 ")
print()

# dd_mooon : Debugging - 플레이어 프린팅 
# print(f"Life Members: {life_members}")  
# print(f"Guest Members: {[p for p in all_players if p not in life_members]}")  

# dd_mooon : Debugging - 리더 프린팅
# print("\nLeaders for all rounds:")
# for round_num, match_type, team in all_rounds_matches:
#     # First player in team is the leader if it's a life member
#     if team and team[0].split('(')[0] in life_members:
#         print(f"Round {round_num}, {match_type}: {team[0]}")
#     else:
#         print(f"Round {round_num}, {match_type}: No leader (no life members in team)")